import openpyxl, json
path = r"D:\Projects\2025\Tata Winger\Drive Upgrade\BOM\BOM Tata Winger.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
out = []
for ws in wb.worksheets:
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            headers.append([None if v is None else str(v) for v in row[:20]])
    out.append({"title": ws.title, "max_row": ws.max_row, "max_col": ws.max_column, "sample": headers[:3]})
print(json.dumps(out, indent=2)[:12000])
